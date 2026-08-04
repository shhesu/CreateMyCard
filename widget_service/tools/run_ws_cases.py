#!/usr/bin/env python3
"""Run every request file in a directory against a WebSocket card-generation API."""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import websockets
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class CaseResult:
    response: dict[str, Any] | None
    error: str | None
    artifact_url: str | None = None
    artifact_markdown: str | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Send every case file verbatim to a WebSocket and save one Excel report."
    )
    parser.add_argument("--ws-url", required=True, help="WebSocket endpoint, e.g. ws://host/path")
    parser.add_argument(
        "--cases-dir", type=Path, required=True, help="Directory containing case files"
    )
    parser.add_argument(
        "--result-dir",
        type=Path,
        help="Excel output directory (default: <cases-dir>/result)",
    )
    parser.add_argument(
        "--case",
        type=Path,
        help="Run only this case file (absolute or relative to --cases-dir)",
    )
    parser.add_argument("--timeout", type=float, default=180.0, help="Per-case timeout in seconds")
    return parser.parse_args()


async def run_case(ws_url: str, payload: str, timeout: float) -> CaseResult:
    """Send one text frame and wait for the final JSON response frame."""
    try:
        async with websockets.connect(ws_url, open_timeout=timeout) as websocket:
            await websocket.send(payload)
            while True:
                raw = await asyncio.wait_for(websocket.recv(), timeout=timeout)
                if isinstance(raw, bytes):
                    raw = raw.decode("utf-8")
                message = json.loads(raw)
                if not isinstance(message, dict):
                    continue
                if message.get("errorCode") not in (None, "0", 0):
                    return CaseResult(message, str(message.get("errorMessage") or "Unknown error"))
                stream_type = (
                    message.get("reply", {}).get("streamInfo", {}).get("streamType")
                )
                if stream_type == "final" or stream_type is None:
                    if message.get("errorCode") in ("0", 0):
                        return CaseResult(message, None)
    except Exception as exc:  # The report must be produced even for transport failures.
        return CaseResult(None, f"{type(exc).__name__}: {exc}")


def extract_artifact_url(response: dict[str, Any]) -> str | None:
    stream_content = (
        response.get("reply", {}).get("streamInfo", {}).get("streamContent", "")
    )
    if not isinstance(stream_content, str):
        return None
    markdown_link = re.search(r"artifactUrl[^\n]*?\]\((https?://[^)]+)\)", stream_content)
    if markdown_link:
        return markdown_link.group(1)
    url = re.search(r"https?://[^\s'\"<>]+", stream_content)
    return url.group(0) if url else None


def download_markdown(url: str, timeout: float) -> str:
    request = urllib.request.Request(url, headers={"User-Agent": "widget-ws-case-runner/1.0"})
    with urllib.request.urlopen(request, timeout=timeout) as response:  # noqa: S310
        return response.read().decode("utf-8")


def extract_code_blocks(markdown: str, language: str) -> list[str]:
    pattern = rf"```{re.escape(language)}\s*\n(.*?)```"
    return re.findall(pattern, markdown, flags=re.IGNORECASE | re.DOTALL)


def extracted_blocks(markdown: str | None, language: str, aliases: tuple[str, ...]) -> str:
    if markdown is None:
        return ""
    return "\n\n".join(
        block.strip()
        for alias in aliases
        for block in extract_code_blocks(markdown, alias)
    )


def write_excel(results: list[tuple[Path, CaseResult]], cases_dir: Path, output_path: Path) -> None:
    """Write every case result into one timestamp-named Excel workbook."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Results"
    headers = [
        "Case",
        "Status",
        "Error",
        "Artifact URL",
        "GenUI",
        "DesignCompactDSL",
        "Response JSON",
    ]
    sheet.append(headers)
    for case_file, result in results:
        sheet.append(
            [
                str(case_file.relative_to(cases_dir)),
                "FAIL" if result.error else "PASS",
                result.error or "",
                result.artifact_url or "",
                extracted_blocks(result.artifact_markdown, "genui", ("genui",)),
                extracted_blocks(
                    result.artifact_markdown,
                    "designcompactdsl",
                    ("designcompactdsl", "design-compact-dsl"),
                ),
                json.dumps(result.response, ensure_ascii=False, indent=2)
                if result.response is not None
                else "",
            ]
        )
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {"A": 32, "B": 10, "C": 42, "D": 60, "E": 60, "F": 60, "G": 60}.items():
        sheet.column_dimensions[column].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


async def main() -> int:
    args = parse_args()
    cases_dir = args.cases_dir.resolve()
    result_dir = (args.result_dir or cases_dir / "result").resolve()
    if not cases_dir.is_dir():
        print(f"cases directory does not exist: {cases_dir}", file=sys.stderr)
        return 2
    if args.case:
        selected_case = args.case if args.case.is_absolute() else cases_dir / args.case
        selected_case = selected_case.resolve()
        if not selected_case.is_file() or cases_dir not in selected_case.parents:
            print(f"case file does not exist in cases directory: {selected_case}", file=sys.stderr)
            return 2
        case_files = [selected_case]
    else:
        case_files = sorted(
            path
            for path in cases_dir.rglob("*")
            if path.is_file() and result_dir not in path.parents
        )
    if not case_files:
        print(f"no case files found in: {cases_dir}", file=sys.stderr)
        return 2

    failures = 0
    results: list[tuple[Path, CaseResult]] = []
    for case_file in case_files:
        result = await run_case(args.ws_url, case_file.read_text(encoding="utf-8"), args.timeout)
        if result.error is None and result.response is not None:
            result.artifact_url = extract_artifact_url(result.response)
            if not result.artifact_url:
                result.error = "successful response does not contain a non-empty artifactUrl"
            else:
                try:
                    result.artifact_markdown = download_markdown(result.artifact_url, args.timeout)
                except Exception as exc:
                    result.error = f"artifact download failed: {type(exc).__name__}: {exc}"
        if result.error:
            failures += 1
        results.append((case_file, result))
        status = "FAIL" if result.error else "PASS"
        print(f"[{status}] {case_file.relative_to(cases_dir)}")
    report_path = result_dir / f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    write_excel(results, cases_dir, report_path)
    print(f"completed {len(case_files)} case(s), failures: {failures}; report: {report_path}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
